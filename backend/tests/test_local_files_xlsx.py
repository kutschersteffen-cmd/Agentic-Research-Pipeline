from openpyxl import Workbook

from arp.ingestion.local_files import parse_file_to_text


def test_parse_xlsx_renders_sheets_as_pipe_delimited_tables(tmp_path):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "EU Taxonomy"
    ws1.append(["KPI", "2025", "2024"])
    ws1.append(["Capex", 6912, 6500])
    ws1.append([None, None, None])  # blank row should be dropped

    wb.create_sheet("Empty Sheet")  # only a header, no data rows

    path = tmp_path / "statbook.xlsx"
    wb.save(path)

    text = parse_file_to_text(path)

    assert "## Sheet: EU Taxonomy" in text
    assert "KPI | 2025 | 2024" in text
    assert "Capex | 6912 | 6500" in text
    assert "Empty Sheet" not in text  # sheet with no data rows contributes nothing


def test_parse_xlsx_drops_trailing_empty_cells(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["A", "B", None, None])

    path = tmp_path / "ragged.xlsx"
    wb.save(path)

    text = parse_file_to_text(path)
    assert "A | B" in text
    assert "A | B | |" not in text
