from __future__ import annotations

import asyncio
import functools
import hashlib
import logging
from importlib.metadata import PackageNotFoundError
from importlib.metadata import version as _pkg_version
from pathlib import Path

from arp.ingestion.base import DocumentSource
from arp.schemas.common import CompanyRef, DocType, SourceDocument
from arp.storage.document_store import DocumentContentStore, derive_doc_id
from arp.storage.safe_path import UnsafeIdentifierError, safe_id

logger = logging.getLogger(__name__)

_TEXT_SUFFIXES = {".txt", ".md"}
_HTML_SUFFIXES = {".html", ".htm"}
_PDF_SUFFIXES = {".pdf"}
_XLSX_SUFFIXES = {".xlsx", ".xlsm"}

# Bump whenever the extraction logic here changes -- including the
# `cursor += len(text) + 2` page-join arithmetic in _extract_pdf_text --
# since that silently changes what a cached page_breaks means without
# changing any package version.
_PARSER_LOGIC_VERSION = 1


@functools.lru_cache(maxsize=1)
def parser_version() -> str:
    """Identifies exactly what produced a DocumentContentStore row, so a
    library upgrade (or a bump to _PARSER_LOGIC_VERSION) orphans old cache
    entries instead of silently mixing parse outputs from two different
    parser behaviors under one key."""
    parts = [f"logic={_PARSER_LOGIC_VERSION}"]
    for pkg in ("pymupdf4llm", "pymupdf", "trafilatura", "openpyxl"):
        try:
            parts.append(f"{pkg}={_pkg_version(pkg)}")
        except PackageNotFoundError:
            parts.append(f"{pkg}=unknown")
    return "|".join(parts)


def _extract_pdf_text(path: Path) -> tuple[str, list[int]]:
    """Renders each page as markdown (via pymupdf4llm/MuPDF) rather than
    plain reading-order text -- disclosure PDFs are table-heavy (segment
    breakdowns, GHG inventories, revenue tables), and MuPDF's table
    detection turns those into real markdown tables instead of pypdf's
    row/column-scrambled flat text. Still zero-LLM, deterministic, and
    fast enough for 4000-company batch runs.

    MuPDF auto-tries an empty user password, so an owner-password-only
    "no printing/copying" disclosure PDF (the common case) reads
    transparently; a genuinely user-password-locked PDF raises, same as
    before, and is caught by the caller as a normal per-file parse error.
    """
    import pymupdf4llm

    pages = pymupdf4llm.to_markdown(str(path), page_chunks=True)
    parts: list[str] = []
    page_breaks: list[int] = []
    cursor = 0
    for page in pages:
        text = page.get("text") or ""
        page_breaks.append(cursor)
        parts.append(text)
        cursor += len(text) + 2  # matches the "\n\n" join below
    return "\n\n".join(parts), page_breaks


def _extract_html_text(path: Path) -> str:
    import trafilatura

    raw = path.read_text(errors="ignore")
    extracted = trafilatura.extract(raw, favor_recall=True)
    if extracted:
        return extracted
    from bs4 import BeautifulSoup

    return BeautifulSoup(raw, "lxml").get_text("\n")


def _extract_xlsx_text(path: Path) -> str:
    """Renders every sheet as a pipe-delimited text table, in document
    order, prefixed by its sheet name -- this is disclosure data (e.g. EU
    Taxonomy revenue/capex/opex tables, GHG footprint statbooks), so
    row/column structure and cross-sheet labels carry the meaning; a naive
    cell dump loses which KPI a number belongs to. Empty rows are skipped;
    trailing empty cells on a row are dropped so ragged tables don't turn
    into walls of pipes."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    sections = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines = [f"## Sheet: {sheet_name}"]
        for row in ws.iter_rows(values_only=True):
            cells = list(row)
            while cells and cells[-1] is None:
                cells.pop()
            if not cells:
                continue
            lines.append(" | ".join("" if c is None else str(c) for c in cells))
        if len(lines) > 1:
            sections.append("\n".join(lines))
    return "\n\n".join(sections)


def parse_file_to_text_with_pages(path: Path) -> tuple[str, list[int]]:
    """Like parse_file_to_text, but also returns PDF page-start char
    offsets (empty for every other format) -- the raw material for
    resolving a grounded citation's exact page number in grounding.py."""
    suffix = path.suffix.lower()
    if suffix in _PDF_SUFFIXES:
        return _extract_pdf_text(path)
    if suffix in _HTML_SUFFIXES:
        return _extract_html_text(path), []
    if suffix in _TEXT_SUFFIXES:
        return path.read_text(errors="ignore"), []
    if suffix in _XLSX_SUFFIXES:
        return _extract_xlsx_text(path), []
    raise ValueError(f"Unsupported document file type: {suffix}")


def parse_file_to_text(path: Path) -> str:
    text, _page_breaks = parse_file_to_text_with_pages(path)
    return text


class LocalFileDocumentSource(DocumentSource):
    """Reads documents from `documents_dir/<company_id>/<doc_type>/*`.

    This is the convention used both by manual user uploads and by the
    discovery crawler's downloader, so anything the crawler finds is
    immediately visible to the extraction pipeline with no extra wiring.
    """

    name = "local_files"

    def __init__(
        self,
        documents_dir: Path,
        content_store: DocumentContentStore | None = None,
        max_concurrent_parses: int = 4,
    ) -> None:
        self.documents_dir = documents_dir
        self._content_store = content_store
        self._parse_sem = asyncio.Semaphore(max_concurrent_parses)

    def _parse_and_identify(
        self, file_path: Path, company_id: str, doc_type: DocType
    ) -> tuple[str | None, str, list[int], str]:
        """Blocking: stat + hash + SQLite + parse, called only via
        asyncio.to_thread so it never stalls the event loop. Returns
        (doc_id, text, page_breaks, text_sha256); doc_id is None when no
        content_store is configured, so SourceDocument falls back to its
        own random default -- exactly today's behavior.
        """
        if self._content_store is not None:
            parsed = self._content_store.get_or_parse(
                file_path, parser_version=parser_version(), parse=parse_file_to_text_with_pages
            )
            doc_id = derive_doc_id(company_id, doc_type.value, parsed.content_key)
            doc_id = self._content_store.register_document(
                doc_id=doc_id,
                company_id=company_id,
                doc_type=doc_type.value,
                content_key=parsed.content_key,
                title=file_path.name,
                local_path=str(file_path),
                source_url=None,
            )
            return doc_id, parsed.full_text, parsed.page_breaks, parsed.text_sha256

        text, page_breaks = parse_file_to_text_with_pages(file_path)
        text_sha256 = hashlib.sha256(text.encode("utf-8")).hexdigest()
        return None, text, page_breaks, text_sha256

    async def fetch(self, company: CompanyRef, doc_types: list[DocType] | None = None) -> list[SourceDocument]:
        try:
            company_dir = self.documents_dir / safe_id(company.company_id, label="company_id")
        except UnsafeIdentifierError:
            # A malformed company_id (e.g. containing "..") has no local
            # documents by definition -- fail this one company closed
            # rather than resolving outside documents_dir, and don't
            # abort the whole batch over one bad universe row.
            logger.warning("Rejected unsafe company_id in local document fetch: %r", company.company_id)
            return []
        if not company_dir.exists():
            return []
        wanted = set(doc_types) if doc_types else None
        file_entries: list[tuple[Path, DocType]] = []
        for doc_type_dir in sorted(company_dir.iterdir()):
            if not doc_type_dir.is_dir():
                continue
            try:
                doc_type = DocType(doc_type_dir.name)
            except ValueError:
                doc_type = DocType.OTHER
            if wanted and doc_type not in wanted:
                continue
            for file_path in sorted(doc_type_dir.iterdir()):
                if file_path.is_file():
                    file_entries.append((file_path, doc_type))

        async def _fetch_one(file_path: Path, doc_type: DocType) -> SourceDocument | None:
            async with self._parse_sem:
                try:
                    doc_id, text, page_breaks, text_sha256 = await asyncio.to_thread(
                        self._parse_and_identify, file_path, company.company_id, doc_type
                    )
                except Exception as exc:  # noqa: BLE001 - isolate one bad file from the whole fetch
                    logger.warning("Failed to parse %s: %s", file_path, exc)
                    return None
            if not text.strip():
                return None
            kwargs: dict = dict(
                company_id=company.company_id,
                doc_type=doc_type,
                title=file_path.name,
                local_path=str(file_path),
                full_text=text,
                sha256=text_sha256,
                page_breaks=page_breaks,
            )
            if doc_id is not None:
                kwargs["doc_id"] = doc_id
            return SourceDocument(**kwargs)

        results = await asyncio.gather(*(_fetch_one(fp, dt) for fp, dt in file_entries))
        return [d for d in results if d is not None]
